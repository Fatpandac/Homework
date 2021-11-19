"""
FileName: 101.py
Author: Fatpandac
Create: 2021/11/19
Description: Reading a excel file called 'data101.xlsx', return a list of sum 
             of each row.
"""


import openpyxl

def main():
    wb = openpyxl.load_workbook('data101.xlsx')
    rows = wb.active.max_row
    cols = wb.active.max_column
    sum_rows_list = []
    for i in range(2, rows+1):
        sum_row_list = []
        for j in range(1, cols+1):
            sum_row_list.append(wb.active.cell(row=i, column=j).value)
        sum_rows_list.append(sum(sum_row_list))

print(main())

