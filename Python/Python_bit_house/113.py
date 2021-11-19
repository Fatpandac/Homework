"""
FileName: 113.py
Author: Fatpandac
Create: 2021/11/20
Description: Replace underline with code.

from openpyxl import load_workbook

def main():
    data = []
    wb = load_workbook('data113.xlsx', _______)
    ws = wb.worksheets[0]
    for index, row in enumerate(ws.rows, start=1):
        if index == 1:
            _______
        data.append(_______)
    return data

print(main())
"""

from openpyxl import load_workbook

def main():
    data = []
    wb = load_workbook('data113.xlsx', data_only=True)
    ws = wb.worksheets[0]
    for index, row in enumerate(ws.rows, start=1):
        if index == 1:
            continue
        data.append(row[3].value)
    return data

print(main())
