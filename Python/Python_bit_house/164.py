"""
FileName: 164.py
Author: Fatpandac
Create: 2021/11/21
Description: Given two string of path and work sheet, return a tuple with 
             row and column.
"""

import openpyxl

def main(workbook_name, worksheet_name):
    wb = openpyxl.load_workbook(workbook_name)
    ws = wb[worksheet_name]
    return (ws.max_row, ws.max_column)
